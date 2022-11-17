from openpyxl import Workbook

wb = Workbook()
#inside the web--> Which sheet we need to work with

#grab the active worksheet
ws = wb.active
ws['A1'] = 'Invoice'
ws['B1'] = 'StockCode'
ws['C1'] = 'Description'
ws['D1'] = 'Quantity'
ws['E1'] = 'InvoiceDate'
ws['F1'] = 'Unitprice'
ws['G1'] = 'Customer'
ws['H1'] = 'Country'
ws['A2'] = 54657655
ws['B2'] = '2638B'
ws['C2'] = 'cvxg'
ws['D2'] = 2
ws['E2'] = '01-01-2022'
ws['F2'] = 2.33
ws['G2'] = 'whitefield'
ws['H2'] = "USA"

#active sheet -- sheet that is open OR mention the sheet 

ws.append([1,2,3])

import datetime

ws['A2'] = datetime.datetime.now()

wb.save("ret.csv")
